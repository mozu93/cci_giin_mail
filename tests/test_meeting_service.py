import openpyxl
from app.database.models import ProcessedAttendanceMail
from app.services.meeting_service import (
    create_meeting, upsert_attendance, export_xlsx, delete_meeting, get_meetings,
    get_attendance_data, get_member_ids_by_status,
    update_actual_status, get_reception_summary, format_minutes_attendee_text,
    export_reception_xlsx,
    build_attendance_export_filename,
)
from app.services.member_service import create_member, delete_member
from app.services.position_service import create_position
from app.services.reception_log_service import create_log, get_logs
from datetime import date, datetime, timedelta

_FUTURE_DATE = date.today() + timedelta(days=30)


def test_build_attendance_export_filename():
    assert build_attendance_export_filename(
        "7月常議員会", "事前", date(2026, 7, 15)
    ) == "20260715_7月常議員会（事前）.xlsx"
    assert build_attendance_export_filename(
        "会議/役員会", "当日", date(2026, 7, 15)
    ) == "20260715_会議＿役員会（当日）.xlsx"


def test_upsert_attendance_saves_notes(db_session):
    member = create_member(db_session, "A-001", "○○商事", "山田太郎")
    meeting = create_meeting(db_session, "常議員会", date(2026, 7, 20))

    record = upsert_attendance(
        db_session, meeting.id, member.id, "出席", notes="体調不良のため途中退席予定")

    assert record.notes == "体調不良のため途中退席予定"


def test_upsert_attendance_notes_defaults_to_empty(db_session):
    member = create_member(db_session, "A-002", "△△産業", "鈴木花子")
    meeting = create_meeting(db_session, "常議員会", date(2026, 7, 20))

    record = upsert_attendance(db_session, meeting.id, member.id, "欠席")

    assert record.notes == ""


def test_export_xlsx_writes_a4_portrait_page_setup(db_session, tmp_path):
    position = create_position(db_session, "議員", 1)
    member = create_member(
        db_session, "A-003", "□□工業", "佐藤次郎",
        title="代表取締役", position_id=position.id)
    meeting = create_meeting(db_session, "定例会議", _FUTURE_DATE)
    upsert_attendance(db_session, meeting.id, member.id, "出席")

    path = tmp_path / "attendance.xlsx"
    export_xlsx(db_session, meeting.id, str(path))

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    assert str(ws.page_setup.paperSize) == str(ws.PAPERSIZE_A4)
    assert ws.page_setup.orientation == "portrait"
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.fitToHeight == 0
    assert ws.sheet_properties.pageSetUpPr.fitToPage is True

    header_row = 4
    assert [ws.cell(row=header_row, column=c).value for c in range(1, 8)] == [
        "No.", "事前", "役職", "事業所名", "所属役職", "氏名", "代理",
    ]
    for c in range(1, 8):
        assert ws.cell(row=header_row, column=c).font.size == 11

    data_row = header_row + 1
    assert ws.cell(row=data_row, column=1).value == 1
    assert ws.cell(row=data_row, column=2).value == "出席"
    assert ws.cell(row=data_row, column=3).value == "議員"
    assert ws.cell(row=data_row, column=4).value == "□□工業"
    assert ws.cell(row=data_row, column=5).value == "代表取締役"
    assert ws.cell(row=data_row, column=6).value == "佐藤次郎"
    assert ws.cell(row=data_row, column=1).font.size == 11
    assert ws.cell(row=data_row, column=1).alignment.shrink_to_fit is True
    assert ws.cell(row=data_row, column=1).alignment.horizontal == "center"
    assert ws.cell(row=data_row, column=2).alignment.horizontal == "center"
    assert ws.cell(row=data_row, column=6).alignment.horizontal == "center"
    assert ws.cell(row=data_row, column=4).alignment.horizontal != "center"

    # 列幅は指定ピクセル値 [30, 45, 235, 129, 93, 45, 141] を
    # (px - 5) / 7 でExcelの文字単位に換算した値と一致すること
    expected_px = [30, 45, 45, 235, 129, 93, 141]
    for c in range(1, 8):
        w = ws.column_dimensions[ws.cell(row=header_row, column=c).column_letter].width
        assert w == round((expected_px[c - 1] - 5) / 7, 2)


def test_delete_meeting_with_reception_logs_does_not_raise(db_session):
    """受付ログ（ReceptionLog）が残っている会議を削除してもFK制約違反で
    クラッシュしないこと（reception_logs.meeting_idの孤立防止の回帰テスト）。"""
    position = create_position(db_session, "議員", 1)
    member = create_member(
        db_session, "A-201", "○○商事", "受付太郎", position_id=position.id)
    meeting = create_meeting(db_session, "定例会議", date(2026, 7, 20))
    upsert_attendance(db_session, meeting.id, member.id, "出席")
    create_log(db_session, meeting.id, member.id, "担当者A", "", "出席")

    delete_meeting(db_session, meeting.id)

    assert get_meetings(db_session) == []
    assert get_logs(db_session, meeting.id) == []


def test_delete_meeting_with_processed_attendance_mail_does_not_raise(db_session):
    """出欠メール取込み済みレコード（ProcessedAttendanceMail）が残っている
    会議を削除してもFK制約違反でクラッシュしないこと。"""
    meeting = create_meeting(db_session, "定例会議", date(2026, 7, 20))
    db_session.add(ProcessedAttendanceMail(
        message_id="msg-1", meeting_id=meeting.id, received_at=datetime.now()))
    db_session.commit()

    delete_meeting(db_session, meeting.id)

    assert get_meetings(db_session) == []
    assert db_session.query(ProcessedAttendanceMail).count() == 0


def test_get_attendance_data_excludes_members_who_joined_after_past_meeting(db_session):
    """開催済みの会議の名簿は固定される：会議日より後に入会した会員は
    一覧に含めない（新しい会員が過去の会議に出席したことにならないように）。"""
    position = create_position(db_session, "議員", 1)
    existing = create_member(
        db_session, "A-501", "既存商事", "既存太郎", position_id=position.id,
        created_at=datetime(2020, 1, 10))
    same_day = create_member(
        db_session, "A-502", "当日商事", "当日花子", position_id=position.id,
        created_at=datetime(2020, 1, 15, 23, 0))
    meeting = create_meeting(db_session, "過去の定例会議", date(2020, 1, 15))
    upsert_attendance(db_session, meeting.id, existing.id, "出席")

    # 会議翌日以降に入会した新規会員
    create_member(
        db_session, "A-503", "新規商事", "新人次郎", position_id=position.id,
        created_at=datetime(2020, 1, 16))

    names = [d["name"] for d in get_attendance_data(db_session, meeting.id)]
    assert "既存太郎" in names
    assert "当日花子" in names  # 会議当日入会は含める
    assert "新人次郎" not in names  # 会議翌日以降の入会は除外


def test_get_attendance_data_keeps_retired_members_with_past_records(db_session):
    """出欠記録が残っている会員は、会議後に退任していても過去の会議の
    出欠一覧から消えないこと（出席実績の保持）。"""
    position = create_position(db_session, "議員", 1)
    attendee = create_member(
        db_session, "A-511", "既存商事", "既存太郎", position_id=position.id,
        created_at=datetime(2020, 1, 1))
    unrelated = create_member(
        db_session, "A-512", "無関係商事", "無関係次郎", position_id=position.id,
        created_at=datetime(2020, 1, 1))
    meeting = create_meeting(db_session, "過去の定例会議", date(2020, 1, 15))
    upsert_attendance(db_session, meeting.id, attendee.id, "出席")

    delete_member(db_session, attendee.id, "担当者A")
    delete_member(db_session, unrelated.id, "担当者A")

    names = [d["name"] for d in get_attendance_data(db_session, meeting.id)]
    assert "既存太郎" in names  # 出欠記録が残っているため退任後も表示
    assert "無関係次郎" not in names  # 出欠記録が無い退任者は表示しない


def test_get_attendance_data_keeps_live_roster_for_upcoming_meeting(db_session):
    """未開催の会議は引き続き最新の名簿をライブ反映する（固定しない）。"""
    position = create_position(db_session, "議員", 1)
    meeting = create_meeting(db_session, "来月の定例会議", date(2099, 1, 15))

    create_member(
        db_session, "A-601", "新規商事", "新人次郎", position_id=position.id)

    names = [d["name"] for d in get_attendance_data(db_session, meeting.id)]
    assert "新人次郎" in names


def test_get_member_ids_by_status_excludes_members_who_joined_after_past_meeting(
        db_session):
    position = create_position(db_session, "議員", 1)
    existing = create_member(
        db_session, "A-701", "既存商事", "既存太郎", position_id=position.id,
        created_at=datetime(2020, 1, 10))
    meeting = create_meeting(db_session, "過去の定例会議", date(2020, 1, 15))
    upsert_attendance(db_session, meeting.id, existing.id, "出席")
    session_actual = db_session
    from app.services.meeting_service import update_actual_status
    update_actual_status(session_actual, meeting.id, existing.id, "出席")

    newcomer = create_member(
        db_session, "A-702", "新規商事", "新人次郎", position_id=position.id,
        created_at=datetime(2020, 1, 16))

    ids = get_member_ids_by_status(db_session, meeting.id, ["未回答"])
    assert newcomer.id not in ids


def test_export_xlsx_attendance_summary_counts(db_session, tmp_path):
    position = create_position(db_session, "議員", 1)
    kanji_position = create_position(db_session, "監事", 2)
    managing_director = create_position(db_session, "専務理事", 3)

    m1 = create_member(db_session, "A-101", "○○商事", "出席太郎",
                       position_id=position.id)
    m2 = create_member(db_session, "A-102", "△△産業", "代理次郎",
                       position_id=position.id)
    m3 = create_member(db_session, "A-103", "□□工業", "委任三郎",
                       position_id=position.id)
    m4 = create_member(db_session, "A-104", "××建設", "欠席四郎",
                       position_id=position.id)
    # 監事は議決権数から除外される
    m5 = create_member(db_session, "A-105", "◇◇興業", "監事五郎",
                       position_id=kanji_position.id)
    # 四日市商工会議所所属でも専務理事は議決権数に含まれる
    m6 = create_member(db_session, "A-106", "四日市商工会議所", "議所六郎",
                       position_id=managing_director.id)

    meeting = create_meeting(db_session, "定例会議", _FUTURE_DATE)
    upsert_attendance(db_session, meeting.id, m1.id, "出席")
    upsert_attendance(db_session, meeting.id, m2.id, "代理")
    upsert_attendance(db_session, meeting.id, m3.id, "委任")
    upsert_attendance(db_session, meeting.id, m4.id, "欠席")
    upsert_attendance(db_session, meeting.id, m5.id, "出席")
    upsert_attendance(db_session, meeting.id, m6.id, "出席")

    path = tmp_path / "attendance_summary.xlsx"
    export_xlsx(db_session, meeting.id, str(path))

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    summary_header_row = next(
        row for row in range(1, ws.max_row + 1)
        if ws.cell(row=row, column=1).value == "出席"
        and ws.cell(row=row, column=2).value == "代理")
    summary_values_row = summary_header_row + 1
    assert [ws.cell(row=summary_header_row, column=c).value for c in range(1, 9)] == [
        "出席", "代理", "委任", "欠席", "議決権数",
        "実出席", "事務局", "合計\n（飲み物用）",
    ]

    # 出席3（m1, m5監事, m6四日市）・代理1・委任1・欠席1
    assert ws.cell(row=summary_values_row, column=1).value == 3
    assert ws.cell(row=summary_values_row, column=2).value == 1
    assert ws.cell(row=summary_values_row, column=3).value == 1
    assert ws.cell(row=summary_values_row, column=4).value == 1
    # 議決権数 = 出席+代理+委任(5) - 監事(m5) = 4（専務理事m6は算入）
    assert ws.cell(row=summary_values_row, column=5).value == 4
    # 実出席者数 = 出席+代理（除外なし）= 4
    assert ws.cell(row=summary_values_row, column=6).value == 4
    # 事務局は空欄で入力可能
    assert ws.cell(row=summary_values_row, column=7).value is None
    # 合計（飲み物用）は実出席者数＋事務局の数式
    assert ws.cell(row=summary_values_row, column=8).value == (
        f"=SUM(F{summary_values_row},G{summary_values_row})")


def test_export_xlsx_summary_note_mentions_exclusion_rule(db_session, tmp_path):
    meeting = create_meeting(db_session, "定例会議", date(2026, 7, 20))
    path = tmp_path / "attendance_note.xlsx"
    export_xlsx(db_session, meeting.id, str(path))

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    note = next(
        ws.cell(row=row, column=1).value
        for row in range(1, ws.max_row + 1)
        if str(ws.cell(row=row, column=1).value or "").startswith("※議決権数"))
    assert "監事" in note
    assert "四日市商工会議所" in note


def test_reception_voting_count_uses_actual_status_and_subtracts_attending_auditors(
        db_session):
    member_position = create_position(db_session, "議員", 1)
    auditor_position = create_position(db_session, "監事", 2)
    attendees = [
        create_member(db_session, "R-001", "A社", "出席者", position_id=member_position.id),
        create_member(db_session, "R-002", "B社", "代理者", position_id=member_position.id),
        create_member(db_session, "R-003", "C社", "委任者", position_id=member_position.id),
        create_member(db_session, "R-004", "D社", "監事出席", position_id=auditor_position.id),
        create_member(db_session, "R-005", "E社", "監事委任", position_id=auditor_position.id),
    ]
    meeting = create_meeting(db_session, "当日会議", _FUTURE_DATE)
    for member, status in zip(
            attendees, ["出席", "代理", "委任", "出席", "委任"]):
        update_actual_status(db_session, meeting.id, member.id, status)

    summary = get_reception_summary(db_session, meeting.id)

    assert summary["出席"] == 2
    assert summary["代理"] == 1
    assert summary["委任"] == 2
    assert summary["監事出席"] == 1
    assert summary["議決権数"] == 4


def test_format_minutes_attendee_text_removes_spaces_and_adds_markers():
    data = [
        {"name": "菅 貴志", "actual_status": "代理"},
        {"name": "川瀬　惠嗣", "actual_status": "委任"},
        {"name": "渡邉 一陽", "actual_status": "出席"},
        {"name": "欠席 太郎", "actual_status": "欠席"},
        {"name": "未受付 花子", "actual_status": ""},
    ]

    assert format_minutes_attendee_text(data) == "菅貴志㈹、川瀬惠嗣(委任)、渡邉一陽"
    assert "\n" not in format_minutes_attendee_text(data)


def test_export_reception_xlsx_contains_actual_status_and_formula(
        db_session, tmp_path):
    position = create_position(db_session, "議員", 1)
    auditor = create_position(db_session, "監事", 2)
    proxy = create_member(
        db_session, "RX-1", "代理会社", "菅 貴志", position_id=position.id)
    delegated = create_member(
        db_session, "RX-2", "委任会社", "川瀬 惠嗣", position_id=position.id)
    attending_auditor = create_member(
        db_session, "RX-3", "監査会社", "監事 太郎", position_id=auditor.id)
    meeting = create_meeting(db_session, "当日受付会議", _FUTURE_DATE)
    update_actual_status(db_session, meeting.id, proxy.id, "代理")
    update_actual_status(db_session, meeting.id, delegated.id, "委任")
    update_actual_status(db_session, meeting.id, attending_auditor.id, "出席")

    path = tmp_path / "reception.xlsx"
    export_reception_xlsx(db_session, meeting.id, str(path))

    wb = openpyxl.load_workbook(path)
    ws = wb["当日受付"]
    summary_header_row = next(
        row for row in range(1, ws.max_row + 1)
        if ws.cell(row=row, column=1).value == "出席"
        and ws.cell(row=row, column=2).value == "代理")
    assert ws.cell(summary_header_row + 1, 7).value == 2
    formula = ws.cell(summary_header_row + 2, 1).value
    assert "代理 1" in formula
    assert "監事の出席者数 1" in formula
    assert "議事録用氏名" not in [
        cell.value for row in ws.iter_rows() for cell in row]
    assert [ws.cell(4, col).value for col in range(1, 7)] == [
        "No.", "当日受付", "事業所名", "会議所役職", "氏名", "代理情報",
    ]
    assert [ws.cell(5 + row, 2).value for row in range(3)] == [
        "代理", "委任", "出席",
    ]
    assert ws.page_setup.orientation == "portrait"
    assert str(ws.page_setup.paperSize) == str(ws.PAPERSIZE_A4)
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.fitToHeight == 0
