import csv
import json
import re
from datetime import date
from sqlalchemy.orm import Session
from app.database.models import Meeting, AttendanceRecord, Member
from app.services.member_service import get_members

STATUS_OPTIONS = ["未回答", "出席", "代理", "委任", "欠席"]


def build_attendance_export_filename(
        meeting_name: str, kind: str, output_date: date | None = None) -> str:
    """Excel出力用の安全な既定ファイル名を作る。"""
    if kind not in ("事前", "当日"):
        raise ValueError("出力種別は「事前」または「当日」を指定してください。")
    safe_name = re.sub(r'[<>:"/\\|?*]', "＿", meeting_name.strip())
    safe_name = safe_name.rstrip(" .") or "会議"
    day = output_date or date.today()
    return f"{day.strftime('%Y%m%d')}_{safe_name}（{kind}）.xlsx"


def create_meeting(session: Session, name: str, meeting_date: date,
                   target_position_ids: list[int] | None = None) -> Meeting:
    ids_json = json.dumps(target_position_ids) if target_position_ids else None
    m = Meeting(name=name, date=meeting_date, target_position_ids=ids_json)
    session.add(m)
    session.commit()
    return m


def get_meetings(session: Session) -> list[Meeting]:
    return session.query(Meeting).order_by(Meeting.date.desc(), Meeting.id.desc()).all()


def delete_meeting(session: Session, meeting_id: int) -> None:
    m = session.get(Meeting, meeting_id)
    if m:
        session.delete(m)
        session.commit()


def upsert_attendance(session: Session, meeting_id: int, member_id: int,
                      status: str, proxy_title: str = "",
                      proxy_name: str = "", notes: str = "") -> AttendanceRecord:
    r = (session.query(AttendanceRecord)
         .filter_by(meeting_id=meeting_id, member_id=member_id)
         .first())
    if r is None:
        r = AttendanceRecord(meeting_id=meeting_id, member_id=member_id)
        session.add(r)
    r.status = status
    r.proxy_title = proxy_title
    r.proxy_name = proxy_name
    r.notes = notes
    session.commit()
    return r


def get_attendance_data(session: Session, meeting_id: int) -> list[dict]:
    """対象会員の出欠データをdictリストで返す（レコード未作成は未回答）。

    会議日を過ぎた会議は名簿を固定する：会議日より後に入会した会員は
    一覧に含めない（新しい会員が過去の会議に出席したことにならないようにする）。
    逆に、会議当時に出欠記録が残っている会員は、その後退任していても
    一覧から消えないようにする（過去の出席実績を保持する）。
    会議日当日までの入会・情報更新は引き続き反映される。"""
    meeting = session.get(Meeting, meeting_id)
    members = get_members(session, active_only=True)
    if meeting and meeting.target_position_ids:
        target_ids = set(json.loads(meeting.target_position_ids))
        members = [m for m in members if m.position_id in target_ids]
    records = {
        r.member_id: r
        for r in session.query(AttendanceRecord)
        .filter_by(meeting_id=meeting_id).all()
    }
    if meeting and is_meeting_past(meeting):
        members = [m for m in members if m.created_at.date() <= meeting.date]
        present_ids = {m.id for m in members}
        retired_ids = set(records.keys()) - present_ids
        if retired_ids:
            members = members + (
                session.query(Member).filter(Member.id.in_(retired_ids)).all())
    result = []
    for m in members:
        r = records.get(m.id)
        result.append({
            "member_id":     m.id,
            "member_number": m.member_number,
            "org_name":      m.organization_name,
            "org_kana":      m.organization_kana or "",
            "title":         m.title or "",
            "name":          m.name,
            "position":      m.position.name if m.position else "",
            "status":        r.status if r else "未回答",
            "actual_status": (r.actual_status or "") if r else "",
            "proxy_title":   r.proxy_title if r else "",
            "proxy_name":    r.proxy_name if r else "",
        })
    return result


def update_actual_status(session: Session, meeting_id: int,
                         member_id: int, actual_status: str) -> None:
    """当日受付ステータスを更新する"""
    r = (session.query(AttendanceRecord)
         .filter_by(meeting_id=meeting_id, member_id=member_id)
         .first())
    if r is None:
        r = AttendanceRecord(meeting_id=meeting_id, member_id=member_id)
        session.add(r)
    r.actual_status = actual_status
    session.commit()


def get_reception_summary(session: Session, meeting_id: int) -> dict:
    """当日受付ステータス（actual_status）による集計"""
    data = get_attendance_data(session, meeting_id)
    counts: dict[str, int] = {"出席": 0, "代理": 0, "委任": 0, "欠席": 0, "未受付": 0}
    for d in data:
        s = d.get("actual_status") or ""
        if s in ("出席", "代理", "委任", "欠席"):
            counts[s] += 1
        else:
            counts["未受付"] += 1
    counts["合計"] = counts["出席"] + counts["代理"] + counts["委任"]
    auditor_attendees = sum(
        1 for d in data
        if d.get("position", "").strip() == "監事"
        and (d.get("actual_status") or "") in ("出席", "代理")
    )
    counts["監事出席"] = auditor_attendees
    counts["議決権数"] = (
        counts["出席"] + counts["代理"] + counts["委任"] - auditor_attendees)
    return counts


def format_minutes_attendee_text(data: list[dict]) -> str:
    """当日受付結果を、議事録へ貼り付ける1行の氏名一覧に変換する。"""
    names = []
    for item in data:
        status = item.get("actual_status") or ""
        if status not in ("出席", "代理", "委任"):
            continue
        name = "".join(str(item.get("name", "")).split())
        if not name:
            continue
        if status == "代理":
            name += "㈹"
        elif status == "委任":
            name += "(委任)"
        names.append(name)
    return "、".join(names)


def export_reception_xlsx(session: Session, meeting_id: int, filepath: str) -> None:
    """当日受付の入力結果を、画面と同じ並び順でExcelへ出力する。"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    meeting = session.get(Meeting, meeting_id)
    data = get_attendance_data(session, meeting_id)
    summary = get_reception_summary(session, meeting_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "当日受付"

    ws["A1"] = meeting.name if meeting else ""
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = meeting.date.strftime("%Y/%m/%d") if meeting else ""
    header_row = 4
    last_data_row = header_row + len(data)
    section_row = last_data_row + 2
    summary_header_row = section_row + 1
    summary_values_row = section_row + 2
    formula_row = section_row + 3
    ws.cell(section_row, 1, "【当日受付集計】").font = Font(bold=True, size=12)

    summary_headers = ["出席", "代理", "委任", "欠席", "未受付", "監事出席", "議決権数"]
    summary_values = [
        summary["出席"], summary["代理"], summary["委任"], summary["欠席"],
        summary["未受付"], summary["監事出席"], summary["議決権数"],
    ]
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, value in enumerate(summary_headers, 1):
        cell = ws.cell(summary_header_row, col, value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        value_cell = ws.cell(summary_values_row, col, summary_values[col - 1])
        value_cell.alignment = Alignment(horizontal="center")
        value_cell.border = border

    formula = (
        f"議決権数 ＝ 当日出席者数（出席 {summary['出席']} ＋ 代理 {summary['代理']}）"
        f" ＋ 委任 {summary['委任']} － 監事の出席者数 {summary['監事出席']}"
        f" ＝ {summary['議決権数']}"
    )
    ws.cell(formula_row, 1, formula)
    ws.merge_cells(start_row=formula_row, start_column=1,
                   end_row=formula_row, end_column=7)

    headers = [
        "No.", "当日受付", "事業所名", "会議所役職", "氏名", "代理情報",
    ]
    for col, value in enumerate(headers, 1):
        cell = ws.cell(header_row, col, value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    for row_no, item in enumerate(data, 1):
        proxy = " ".join(
            value for value in (item.get("proxy_title", ""), item.get("proxy_name", ""))
            if value)
        values = [
            row_no, item.get("actual_status") or "", item["org_name"],
            item["position"], item["name"], proxy,
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(header_row + row_no, col, value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", shrink_to_fit=True)

    widths = [6, 12, 34, 16, 16, 24, 12]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    last_row = formula_row
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:F{last_row}"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.print_area = f"A1:G{last_row}"
    wb.save(filepath)


def get_summary(session: Session, meeting_id: int) -> dict:
    data = get_attendance_data(session, meeting_id)
    counts: dict[str, int] = {"出席": 0, "代理": 0, "委任": 0, "欠席": 0, "未回答": 0}
    for d in data:
        s = d["status"]
        counts[s] = counts.get(s, 0) + 1
    counts["合計"] = counts["出席"] + counts["代理"] + counts["委任"]
    return counts


def is_meeting_past(meeting: Meeting) -> bool:
    """会議日を過ぎている（開催済み）かどうかを返す"""
    return date.today() > meeting.date


def get_member_ids_by_status(session: Session, meeting_id: int,
                             statuses: list[str]) -> set[int]:
    """指定した会議・ステータスに該当する会員IDのセットを返す。
    「未回答」は事前登録（status）、それ以外（出席・代理・委任・欠席）は
    当日受付結果（actual_status）で判定する（会議日の前後は問わない）。"""
    meeting = session.get(Meeting, meeting_id)
    if not meeting:
        return set()
    members = get_members(session, active_only=True)
    if meeting.target_position_ids:
        target_ids = set(json.loads(meeting.target_position_ids))
        members = [m for m in members if m.position_id in target_ids]
    if is_meeting_past(meeting):
        members = [m for m in members if m.created_at.date() <= meeting.date]
    records = {
        r.member_id: r
        for r in session.query(AttendanceRecord)
        .filter_by(meeting_id=meeting_id).all()
    }
    result_ids = set()
    other_statuses = [s for s in statuses if s != "未回答"]
    for m in members:
        r = records.get(m.id)
        if "未回答" in statuses:
            pre_status = r.status if r else "未回答"
            if pre_status == "未回答":
                result_ids.add(m.id)
                continue
        if other_statuses:
            actual_status = r.actual_status if r and r.actual_status else ""
            if actual_status in other_statuses:
                result_ids.add(m.id)
    return result_ids


def export_csv(session: Session, meeting_id: int, filepath: str) -> None:
    meeting = session.get(Meeting, meeting_id)
    data = get_attendance_data(session, meeting_id)
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["会議名", meeting.name if meeting else ""])
        writer.writerow(["会員番号", "事業所名", "会議所役職", "氏名",
                         "ステータス", "代理役職名", "代理氏名"])
        for d in data:
            writer.writerow([
                d["member_number"], d["org_name"], d["position"], d["name"],
                d["status"], d["proxy_title"], d["proxy_name"],
            ])


_XLSX_HEADERS = ["No.", "事前", "役職", "事業所名", "所属役職", "氏名", "代理"]
_XLSX_FONT_SIZE = 11
_XLSX_CENTER_COLUMNS = {1, 2, 6}  # No., 事前, 氏名
# 列幅（ピクセル指定）。Excelの列幅（文字単位）へは (px - 5) / 7 で換算する
# （既定フォントCalibri 11・既定列幅8.43文字=64pxを基準とした変換式）。
_XLSX_COLUMN_WIDTHS_PX = [30, 45, 45, 235, 129, 93, 141]

# 議決権数の集計から除外する会議所役職・事業所キーワード
_VOTING_EXCLUDED_POSITION = "監事"
_VOTING_EXCLUDED_ORG_KEYWORD = "四日市商工会議所"

_SUMMARY_HEADERS = ["出席", "代理", "委任", "欠席", "議決権数",
                    "実出席", "事務局", "合計\n（飲み物用）"]
_SUMMARY_COLUMN_WIDTHS_PX = [60, 60, 60, 60, 80, 100, 60, 80]
_SUMMARY_OFFICE_COL = 7   # 事務局（空欄入力用）
_SUMMARY_TOTAL_COL = 8    # 合計（飲み物用、数式）
_SUMMARY_ACTUAL_COL = 6   # 実出席者数


def _px_to_excel_width(px: int) -> float:
    return round((px - 5) / 7, 2)


def _calc_attendance_summary(data: list[dict]) -> dict:
    """出欠状況・議決権数・実出席者数を集計する。

    議決権数は出席・代理・委任の合計から、会議所役職が「監事」の会員と、
    事業所名に「四日市商工会議所」を含む会員を除いた人数。
    ただし「専務理事」は四日市商工会議所所属でも議決権数に含める。
    実出席（飲み物注文用）は出席・代理の合計（除外なし）。
    """
    counts = {"出席": 0, "代理": 0, "委任": 0, "欠席": 0}
    voting_count = 0
    actual_count = 0
    for d in data:
        status = d["status"]
        if status in counts:
            counts[status] += 1
        position = d["position"].strip()
        excluded = (
            position == _VOTING_EXCLUDED_POSITION
            or (
                _VOTING_EXCLUDED_ORG_KEYWORD in d["org_name"]
                and position != "専務理事"
            )
        )
        if status in ("出席", "代理", "委任") and not excluded:
            voting_count += 1
        if status in ("出席", "代理"):
            actual_count += 1
    counts["議決権数"] = voting_count
    counts["実出席者数"] = actual_count
    return counts


def export_xlsx(session: Session, meeting_id: int, filepath: str) -> None:
    """会議の出欠一覧をA4縦向き印刷向けに整形したExcelファイルに書き出す。
    行順は会員一覧の並び順（会議所役職順）に従う。行数が多い場合は
    自動的に複数ページに分かれて印刷される。"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    meeting = session.get(Meeting, meeting_id)
    data = get_attendance_data(session, meeting_id)
    summary = _calc_attendance_summary(data)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "出欠一覧"

    n_cols = max(len(_XLSX_HEADERS), len(_SUMMARY_HEADERS))
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(bold=True, color="FFFFFF", size=_XLSX_FONT_SIZE)
    data_font = Font(size=_XLSX_FONT_SIZE)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    input_fill = PatternFill("solid", fgColor="FFF2CC")

    # 行番号は固定レイアウト（openpyxlのws.append()は空行だとmax_rowが
    # 進まないため、以降すべて行番号を明示したcell()で書き込む）。
    ws.cell(row=1, column=1, value=meeting.name if meeting else "").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1,
            value=meeting.date.strftime("%Y/%m/%d") if meeting else "").font = Font(
        size=11, color="666666")

    header_row = 4
    last_data_row = header_row + len(data)

    # --- 名簿下部の出欠状況集計 ---
    section_row = last_data_row + 2
    ws.cell(row=section_row, column=1, value="【出欠状況集計】").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=section_row, start_column=1,
                   end_row=section_row, end_column=n_cols)

    summary_header_row = section_row + 1
    for col, text in enumerate(_SUMMARY_HEADERS, start=1):
        c = ws.cell(row=summary_header_row, column=col, value=text)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    summary_values_row = section_row + 2
    summary_values = [
        summary["出席"], summary["代理"], summary["委任"], summary["欠席"],
        summary["議決権数"], summary["実出席者数"], None, None,
    ]
    for col, val in enumerate(summary_values, start=1):
        c = ws.cell(row=summary_values_row, column=col, value=val)
        c.font = (Font(bold=True, size=_XLSX_FONT_SIZE)
                  if col == _SUMMARY_TOTAL_COL else data_font)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.cell(row=summary_values_row, column=_SUMMARY_OFFICE_COL).fill = input_fill
    ws.cell(row=summary_values_row, column=_SUMMARY_TOTAL_COL).value = (
        f"=SUM({get_column_letter(_SUMMARY_ACTUAL_COL)}{summary_values_row},"
        f"{get_column_letter(_SUMMARY_OFFICE_COL)}{summary_values_row})"
    )

    note_row = section_row + 3
    ws.cell(row=note_row, column=1,
            value="※議決権数は出席・代理・委任の合計から、監事および四日市商工会議所を除いた人数です。"
                  "ただし、専務理事は議決権数に含みます。"
                  "事務局欄に人数を入力すると合計（飲み物用）が自動計算されます。").font = Font(
        size=9, color="666666")
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=n_cols)

    for col, text in enumerate(_XLSX_HEADERS, start=1):
        c = ws.cell(row=header_row, column=col, value=text)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    for i, d in enumerate(data, start=1):
        proxy_info = " ".join(p for p in [d["proxy_title"], d["proxy_name"]] if p)
        values = [i, d["status"], d["position"], d["org_name"], d["title"],
                  d["name"], proxy_info]
        for col, val in enumerate(values, start=1):
            ws.cell(row=header_row + i, column=col, value=val)

    last_row = note_row
    for row in ws.iter_rows(min_row=header_row + 1, max_row=last_data_row,
                            max_col=len(_XLSX_HEADERS)):
        for cell in row:
            cell.border = border
            cell.font = data_font
            cell.alignment = Alignment(vertical="center", shrink_to_fit=True)
    for row_idx in range(header_row + 1, last_data_row + 1):
        for col in _XLSX_CENTER_COLUMNS:
            ws.cell(row=row_idx, column=col).alignment = Alignment(
                horizontal="center", vertical="center", shrink_to_fit=True)

    # 列幅は指定ピクセル値で固定し、はみ出す分はセル書式の
    # 「縮小して全体を表示」でフォントサイズを自動調整させる。
    for col, px in enumerate(_XLSX_COLUMN_WIDTHS_PX, start=1):
        ws.column_dimensions[get_column_letter(col)].width = _px_to_excel_width(px)
    # 出欠状況集計のみで使う8列目（合計）の幅
    for col in range(len(_XLSX_COLUMN_WIDTHS_PX) + 1, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = _px_to_excel_width(
            _SUMMARY_COLUMN_WIDTHS_PX[col - 1])

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # A4縦向き印刷設定：横幅は1ページに収め、行数が多い場合は縦方向に複数ページへ分割
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.print_area = f"A1:{get_column_letter(n_cols)}{last_row}"
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6
    ws.print_options.horizontalCentered = True

    wb.save(filepath)
